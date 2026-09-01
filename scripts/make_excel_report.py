"""Turn out/batch_test_results.json into a formatted Excel workbook:
one row per test run, plus a summary sheet.
"""
from __future__ import annotations

import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
RESULTS_PATH = ROOT / "out" / "batch_test_results.json"
XLSX_PATH = ROOT / "out" / "batch_test_results.xlsx"

COLUMNS = [
    ("test_id", "Test ID", 22),
    ("input_image", "Input Image", 30),
    ("status", "Status", 20),
    ("faces_detected", "Faces Detected", 14),
    ("detect_score", "Detect Score", 13),
    ("total_visual_matches", "Total Lens Matches", 17),
    ("social_candidates", "Social Candidates", 16),
    ("candidates_checked", "Candidates Checked", 17),
    ("best_cosine", "Best Cosine", 12),
    ("matched", "Matched?", 10),
    ("matched_source", "Matched Platform", 16),
    ("matched_url", "Matched Post URL", 45),
    ("record_hash", "Record Hash (keccak256)", 45),
    ("chain_mode", "Chain", 8),
    ("anchor_tx", "Anchor Tx Hash", 45),
    ("anchor_verified", "Anchor Verified?", 15),
    ("elapsed_seconds", "Elapsed (s)", 11),
    ("error", "Error", 40),
]

STATUS_COLORS = {
    "OK": "C6EFCE",
    "NO_MATCH": "FFEB9C",
    "NO_SOCIAL_CANDIDATES": "FFEB9C",
    "NO_FACE": "FFC7CE",
    "ERROR": "FFC7CE",
}
STATUS_FONT = {
    "OK": "006100",
    "NO_MATCH": "9C6500",
    "NO_SOCIAL_CANDIDATES": "9C6500",
    "NO_FACE": "9C0006",
    "ERROR": "9C0006",
}


def main() -> None:
    results = json.loads(RESULTS_PATH.read_text())

    wb = Workbook()

    # --- Results sheet -------------------------------------------------
    ws = wb.active
    ws.title = "Test Results"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (_key, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(results, 2):
        for col_idx, (key, _label, _width) in enumerate(COLUMNS, 1):
            value = r.get(key, "")
            if key == "matched":
                value = "Yes" if value else "No"
            if key == "anchor_verified":
                value = "Yes" if value else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(key in ("matched_url", "record_hash", "anchor_tx", "error")))

        status = r.get("status", "")
        fill_color = STATUS_COLORS.get(status)
        font_color = STATUS_FONT.get(status)
        if fill_color:
            status_col = [i for i, (k, _, _) in enumerate(COLUMNS, 1) if k == "status"][0]
            c = ws.cell(row=row_idx, column=status_col)
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            c.font = Font(color=font_color, bold=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(results) + 1}"

    # --- Summary sheet ---------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    ok = sum(1 for r in results if r["status"] == "OK")
    no_match = sum(1 for r in results if r["status"] == "NO_MATCH")
    no_social = sum(1 for r in results if r["status"] == "NO_SOCIAL_CANDIDATES")
    no_face = sum(1 for r in results if r["status"] == "NO_FACE")
    error = sum(1 for r in results if r["status"] == "ERROR")
    cosines = [r["best_cosine"] for r in results if r.get("best_cosine") is not None]
    avg_cosine = round(sum(cosines) / len(cosines), 4) if cosines else None
    anchored = sum(1 for r in results if r.get("anchor_verified"))

    summary_rows = [
        ("Total tests run", total),
        ("Matched + anchored on-chain (OK)", ok),
        ("Searched OK, no verified match (NO_MATCH)", no_match),
        ("Searched OK, no social-media candidates (NO_SOCIAL_CANDIDATES)", no_social),
        ("No face detected in input (NO_FACE)", no_face),
        ("Errors (network/API) (ERROR)", error),
        ("Match rate", f"{ok}/{total} ({round(100*ok/total,1) if total else 0}%)"),
        ("Average best cosine similarity across all tests", avg_cosine),
        ("On-chain anchors independently re-verified", f"{anchored}/{ok}"),
        ("Chain used", results[0]["chain_mode"] if results else ""),
    ]
    ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 25
    for i, (label, value) in enumerate(summary_rows, 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH} ({total} test rows)")


if __name__ == "__main__":
    main()
